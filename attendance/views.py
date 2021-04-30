from rest_framework import generics, views, viewsets
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from .models import AttendanceStatus, TeachersAttendance
from .serializers import (
	AttendanceStatusSerializer, TeachersAttendanceSerializer)

class AttendanceStatusViewSet(viewsets.ModelViewSet):
	queryset = AttendanceStatus.objects.all()
	serializer_class = AttendanceStatusSerializer

class TeachersAttendanceViewSet(viewsets.ModelViewSet):
	queryset = TeachersAttendance.objects.all()
	serializer_class = TeachersAttendanceSerializer

class TeachersAttendanceListView(views.APIView):
	"""
    List all students, or create a new student.
    """
	def get(self, request, format=None):
		attendances = TeachersAttendance.objects.all()
		serializer = TeachersAttendanceSerializer(attendances, many=True)
		return Response(serializer.data)

	def post(self, request, format=None):
		serializer = TeachersAttendanceSerializer(data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data, status=status.HTTP_201_CREATED)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TeachersAttendanceDetailView(views.APIView):
	def get_object(self, pk):
		try:
			return TeachersAttendance.objects.get(pk=pk)
		except TeachersAttendance.DoesNotExist:
			raise Http404
	def get(self, request, pk, format=None):
		attendance = self.get_object(pk)
		serializer = TeachersAttendanceSerializer(attendance)
		return Response(serializer.data)
		
	def put(self, request, pk, format=None):
		attendance = self.get_object(pk)
		serializer = TeachersAttendanceSerializer(attendance, data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
	
	def delete(self, request, pk, format=None):
		attendance = self.get_object(pk)
		attendance.delete()
		return Response(status=status.HTTP_204_NO_CONTENT)


class TeachersAttendanceBulkCreateView(generics.CreateAPIView):
	serializer_class = TeachersAttendanceSerializer
	queryset = TeachersAttendance.objects.all()

	def post(self, request):
		print(request.data)
		serializer = TeachersAttendanceSerializer(request.data)
		#serializer.is_valid()
		teachers_attendance = serializer.create(request)
		if teachers_attendance:
			return Response(status=status.HTTP_201_CREATED)
		return Response(status=status.HTTP_400_BAD_REQUEST)

class TeachersAttendanceBulkUploadView(views.APIView):
	"""
	This uploads bulk daily teacher's attendance from an excel file
	"""

	parser_class = [FileUploadParser]
	def post(self, request, filename, format="xlsx"):
		file_obj = request.data
		xlfile = file_obj["filename"]

		#print(xlfile)
		wb = load_workbook(xlfile)
		ws = wb.active
		#print(ws.title)

		studentz = []
		for row in ws.iter_rows(min_row=2, max_col=9, max_row=12, values_only=True):
			studentz.append(row)
			#print(api)
			
		students = []
		for i in range(len(studentz)):
			student = {
				"date": f"{studentz[i][0]}",
				"time_in": f"{studentz[i][1]}",
				"time_out": f"{studentz[i][2]}",
				"teacher": f"{studentz[i][3]}",
				"status": f"{studentz[i][4]}",
					}
			students.append(student)
		
		for student in students:
			if student in Student.objects.all():
				print("student exists!")
				continue
			else: 
				serializer = StudentSerializer(data=student)
				if serializer.is_valid():
					serializer.save()
					#return Response(serializer.data, status=status.HTTP_201_CREATED)
				#return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


		#print(students)


		student = Student()


		return Response(status=204)


