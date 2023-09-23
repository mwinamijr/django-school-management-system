from rest_framework import generics, views
from rest_framework.parsers import FileUploadParser
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework import status
from django.http import Http404
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from academic.models import Student

from .serializers import FileUploadSerializer
from academic.serializers import StudentSerializer

class StudentListView(views.APIView):
	"""
    List all students, or create a new student.
    """
	#permission_classes = [IsAuthenticated]
	def get(self, request, format=None):
		students = Student.objects.all()
		serializer = StudentSerializer(students, many=True)
		return Response(serializer.data)

	def post(self, request, format=None):
		serializer = StudentSerializer(data=request.data)

		print(serializer.is_valid())
		print(request.data)
		if serializer.is_valid():
			student = serializer.create(request)
			if student:
				#serializer.save()
				return Response(status=status.HTTP_201_CREATED)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class StudentDetailView(views.APIView):
	permission_classes = [IsAuthenticated]
	def get_object(self, pk):
		try:
			return Student.objects.get(pk=pk)
		except Student.DoesNotExist:
			raise Http404
	def get(self, request, pk, format=None):
		student = self.get_object(pk)
		serializer = StudentSerializer(student)
		return Response(serializer.data)
		
	def put(self, request, pk, format=None):
		student = self.get_object(pk)
		serializer = StudentSerializer(student, data=request.data)
		if serializer.is_valid():
			serializer.save()
			return Response(serializer.data)
		return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
	
	def delete(self, request, pk, format=None):
		student = self.get_object(pk)
		student.delete()
		return Response(status=status.HTTP_204_NO_CONTENT)

class StudentBulkUploadView(views.APIView):
	"""
	This uploads bulk number of students from excel file
	"""

	parser_class = [FileUploadParser]
	def post(self, request, filename, format="xlsx"):
		file_obj = request.data
		print(request.data)
		
		xlfile = file_obj["filename"]

		wb = load_workbook(xlfile)
		ws = wb.active
		print(ws.title)
		
		studentz = []
		for row in ws.iter_rows(min_row=2, max_col=9, max_row=6, values_only=True):
			studentz.append(row)
		#print(studentz)
		
		students = []
		for i in range(len(studentz)):
			student = {
				"addmission_number": f"{studentz[i][0]}",
				"first_name": f"{studentz[i][1]}",
				"middle_name": f"{studentz[i][2]}",
				"last_name": f"{studentz[i][3]}",
				"grad_date": f"{studentz[i][4]}",
				"sex": f"{studentz[i][5]}",
				"birthday": f"{studentz[i][6]}",
				"class_level": f"{studentz[i][8]}",	
					}
			students.append(student)
		
		for student in students:
			if student in Student.objects.all():
				print("student exists!")
				continue
			else: 
				serializer = StudentSerializer(data=student)
				if serializer.is_valid():
					serializer.save()

		return Response(status=status.HTTP_201_CREATED)



'''
class StudentHealthRecordViewSet(viewsets.ModelViewSet):
	queryset = StudentHealthRecord.objects.all()
	serializer_class = StudentHealthRecordSerializer

class GradeScaleViewSet(viewsets.ModelViewSet):
	queryset = GradeScale.objects.all()
	serializer_class = GradeScaleSerializer

class GradeScaleRuleViewSet(viewsets.ModelViewSet):
	queryset = GradeScaleRule.objects.all()
	serializer_class = GradeScaleRuleSerializer

class SchoolYearViewSet(viewsets.ModelViewSet):
	queryset = SchoolYear.objects.all()
	serializer_class = SchoolYearSerializer

class MessageToStudentViewSet(viewsets.ModelViewSet):
	queryset = MessageToStudent.objects.all()
	serializer_class = MessageToStudentSerializer
'''
